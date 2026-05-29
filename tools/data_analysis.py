"""Data analysis tools: Excel/CSV reading, chart generation, code execution."""
import csv
import io
import json
import os
import re
import sys
from pathlib import Path

from tools._shared import emit_card


TOOLS = [
    {"name": "analyze_spreadsheet",
     "description": "Apre file Excel (xlsx) o CSV, restituisce headers + prime righe + statistiche numeriche.",
     "input_schema": {"type": "object", "properties": {
         "path": {"type": "string"},
         "sheet": {"type": "string", "description": "Nome foglio xlsx (default: primo)"},
         "max_rows": {"type": "integer"},
     }, "required": ["path"]}},
    {"name": "make_chart",
     "description": "Genera un grafico (linee/barre) da dati x,y. Mostra una card.",
     "input_schema": {"type": "object", "properties": {
         "title": {"type": "string"},
         "x_label": {"type": "string"},
         "y_label": {"type": "string"},
         "data": {"type": "array", "items": {"type": "object", "properties": {
             "label": {"type": "string"},
             "value": {"type": "number"},
         }, "required": ["label", "value"]}},
         "type": {"type": "string", "enum": ["bar", "line"]},
     }, "required": ["data"]}},
    {"name": "code_exec",
     "description": "Esegue codice Python in sandbox (no rete, no filesystem read). Utile per calcoli avanzati. Es: 'sum([1,2,3])', 'sorted([5,3,1])'.",
     "input_schema": {"type": "object", "properties": {
         "code": {"type": "string", "description": "Codice Python. Usa 'print()' per output."},
     }, "required": ["code"]}},
]


def _read_csv(path: str, max_rows: int = 100):
    rows = []
    headers = []
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except Exception:
                dialect = csv.excel
            reader = csv.reader(f, dialect)
            for i, row in enumerate(reader):
                if i == 0:
                    headers = row
                else:
                    rows.append(row)
                if i > max_rows:
                    break
    except Exception as e:
        return None, str(e)
    return (headers, rows), None


def _read_xlsx(path: str, sheet: str = "", max_rows: int = 100):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, "openpyxl non disponibile"
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter, []))
        rows = []
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                break
            rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
        return (headers, rows), None
    except Exception as e:
        return None, str(e)


def _stats_for_columns(headers, rows):
    """Compute basic stats per column (count, min/max/avg if numeric)."""
    n_cols = len(headers)
    stats = []
    for ci in range(n_cols):
        col = [r[ci] if ci < len(r) else "" for r in rows]
        nums = []
        for v in col:
            try:
                # Try to parse as number (also commas as decimal separator)
                s = str(v).replace(",", ".").strip()
                if s:
                    nums.append(float(s))
            except Exception:
                pass
        if nums and len(nums) >= len(col) * 0.5:
            stats.append({
                "name": headers[ci] if ci < len(headers) else f"col{ci}",
                "type": "numeric",
                "count": len(nums),
                "min": round(min(nums), 2),
                "max": round(max(nums), 2),
                "avg": round(sum(nums) / len(nums), 2),
                "sum": round(sum(nums), 2),
            })
        else:
            stats.append({
                "name": headers[ci] if ci < len(headers) else f"col{ci}",
                "type": "text",
                "count": sum(1 for x in col if x),
                "unique": len(set(col)),
            })
    return stats


def _safe_exec(code: str):
    """Execute Python in a restricted sandbox. Capture stdout."""
    import builtins
    safe_builtins = {
        "abs": abs, "all": all, "any": any, "bin": bin, "bool": bool,
        "chr": chr, "dict": dict, "divmod": divmod, "enumerate": enumerate,
        "filter": filter, "float": float, "hex": hex, "int": int,
        "isinstance": isinstance, "issubclass": issubclass, "iter": iter,
        "len": len, "list": list, "map": map, "max": max, "min": min,
        "next": next, "oct": oct, "ord": ord, "pow": pow, "print": print,
        "range": range, "repr": repr, "reversed": reversed, "round": round,
        "set": set, "slice": slice, "sorted": sorted, "str": str, "sum": sum,
        "tuple": tuple, "type": type, "zip": zip, "True": True, "False": False, "None": None,
    }
    import math, statistics, random, datetime, json as _json, re as _re
    safe_globals = {
        "__builtins__": safe_builtins,
        "math": math, "statistics": statistics, "random": random,
        "datetime": datetime, "json": _json, "re": _re,
    }
    buf = io.StringIO()
    old_stdout = sys.stdout
    sys.stdout = buf
    try:
        # First try eval (single expression) for convenience
        try:
            result = eval(code, safe_globals)
            sys.stdout = old_stdout
            output = buf.getvalue()
            if result is not None and not output:
                return f"= {result!r}"
            return output + (f"\n= {result!r}" if result is not None else "")
        except SyntaxError:
            # Fall back to exec for multi-statement code
            exec(code, safe_globals)
            sys.stdout = old_stdout
            return buf.getvalue() or "(no output)"
    except Exception as e:
        sys.stdout = old_stdout
        return f"Errore: {type(e).__name__}: {e}"


def run(name, args):
    if name == "analyze_spreadsheet":
        path = args.get("path", "").strip()
        if not os.path.exists(path):
            return f"File non trovato: {path}"
        sheet = args.get("sheet", "")
        max_rows = int(args.get("max_rows", 50))

        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            result, err = _read_csv(path, max_rows)
        elif ext in (".xlsx", ".xlsm"):
            result, err = _read_xlsx(path, sheet, max_rows)
        else:
            return "Formato non supportato. Solo CSV/XLSX."

        if err or not result:
            return f"Errore lettura file: {err}"

        headers, rows = result
        stats = _stats_for_columns(headers, rows)

        out = [f"File: {os.path.basename(path)}",
               f"Colonne: {len(headers)}", f"Righe lette: {len(rows)}", ""]
        out.append("ANTEPRIMA (prime 5 righe):")
        out.append(" | ".join(str(h)[:20] for h in headers))
        for row in rows[:5]:
            out.append(" | ".join(str(c)[:20] for c in row))
        out.append("")
        out.append("STATISTICHE COLONNE:")
        for s in stats:
            if s["type"] == "numeric":
                out.append(f"  {s['name']}: numerico, {s['count']} valori, min={s['min']}, max={s['max']}, media={s['avg']}, somma={s['sum']}")
            else:
                out.append(f"  {s['name']}: testo, {s['count']} valori, {s['unique']} unici")
        return "\n".join(out)

    if name == "make_chart":
        data = args.get("data", [])
        if not data:
            return "Dati richiesti."
        title = args.get("title", "Grafico")
        chart_type = args.get("type", "bar")
        emit_card("chart", {
            "title": title,
            "type": chart_type,
            "x_label": args.get("x_label", ""),
            "y_label": args.get("y_label", ""),
            "data": data,
        })
        return f"Grafico '{title}' generato ({len(data)} punti)."

    if name == "code_exec":
        code = args.get("code", "")
        if not code:
            return "Codice vuoto."
        return _safe_exec(code)

    return "?"
